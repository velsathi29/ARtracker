from django.shortcuts import render
from django.views.generic import TemplateView
from django.views.generic import ListView
from .forms import FileUploadForm, MyModelForm
from django.shortcuts import redirect
from sqlalchemy import create_engine
from.models import Detail
import json

from django.core.paginator import Paginator
from django.core.serializers import serialize
from django.http import JsonResponse
from django.views import View
from datetime import datetime

# Create your views here.
from django.conf import settings
import pandas as pd

#to import data from excel to DB
from http.client import HTTPResponse
from django.urls import reverse
import pandas as pd
import os
import datetime
from django.core.files.storage import FileSystemStorage


#  /index upload page
from django.http import HttpResponse,HttpResponseRedirect  
from ardailytracker.functions import handle_uploaded_file  
from ardailytracker.forms import Det_Form
import openpyxl
from django.template import loader
from django.db.models import Q #useful for OR condition



#export to excel
import xlwt
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
# from django.contrib.auth.models import Detail



# for restricting this url
def prem(request):
    current_user = request.user
    top_users=['indira','premkumar','PREMKUMAR','shalini']
    #if current_user in top_users:
    mydata = Detail.objects.all() #to order based on status
    # for data in mydata:
    #     data['aging_Lag'] = datetime.datetime.strptime(member['aging_Lag'], '%m %H:%M:%S.%f').strftime('%Y-%m-%d %H:%M:%S')
    template = loader.get_template('template1.html') #template.html(previously)
    context = {
    'mymembers': mydata,
        }
    return HttpResponse(template.render(context, request))
    



@login_required
def testpurp(request):
    current_user = request.user
    top_users=['indira','premkumar','PREMKUMAR','shalini']
    #if current_user in top_users:
    mydata = Detail.objects.all() 
    template = loader.get_template('testpurp.html') #template.html(previously)
    context = {
    'mymembers': mydata,
        }
    return HttpResponse(template.render(context, request))
    



# for viewing Detail model as a table (view detail/)
def tablemodel(request):
    current_user = request.user
    top_users=['indira','premkumar','PREMKUMAR','shalini']
    #if current_user in top_users:
    mydata = Detail.objects.all() 
    template = loader.get_template('detailmodel.html')
    context = {
    'mymembers': mydata,
        }
    return HttpResponse(template.render(context, request))




    
    




@login_required
def excellikefilter(request):
    
    # render function takes argument  - request
    # and return HTML as response
    return render(request, "excellikefilter.html")




#FOR EXPORTING DATABASE AS EXCEL FILE
def export_users_xlsx(request):

    current_datetime = datetime.datetime.now().strftime("%m/%d/%Y %I:%M:%S %p")
    file_name='ARDailyTracker'+str(current_datetime)+'.xls'

    response=HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition']='attachment;filename="ARDailyTracker.xls"'
    #response['Content-Disposition']='attachment;filename="ARDailyTracker.xls"'

    wb=xlwt.Workbook(encoding='utf-8')
    ws=wb.add_sheet('Tasks')

    row_num=0

    font_style=xlwt.XFStyle()
    font_style.font.bold=True

    columns=['Received','Service Date','Client Name','Claim Status','Claim Number','Payer','Charge $','Balance $','Upload Date','Assigned To','Allocation Date','Call Type','Disposition','Denial Code','Step','Next Best Action For Success Rate','Follow Up Date','Follow Up Comments','Final Status','Worked By','Worked On']
    for col_num in range(len(columns)):

        ws.write(row_num,col_num,columns[col_num],font_style)

    font_style=xlwt.XFStyle()

    rows=Detail.objects.all().values_list('received','service_Date','client','claim_Status','account_or_Claim_Number','payer','charge_USD','balance_USD','upload_Date','assigned_To','allocation_Date','call_Type','disposition','denial_Code','step','next_Best_Action_For_Success_Rate','follow_up_Date','follow_up_Comments','final_Status','worked_By','worked_On')
    
    for row in rows:
        row_num+=1
        for col_num in range(len(row)):
            if isinstance(row[col_num], datetime.datetime):

                date_time = row[col_num].strftime('%m-%d-%Y %H:%M:%S')
                ws.write(row_num, col_num, date_time, font_style)
            else:
                ws.write(row_num,col_num,row[col_num],font_style)
    
    wb.save(response)
    return response






# TO GET FORM (UPDATE.HTML) TEMPLATE AFTER CLICKING UNIQUE ID 
@login_required
def update(request, id):
  mymember = Detail.objects.get(id=id)
  
  template = loader.get_template('update.html')
  context = {
    'mymember': mymember,
  }
  return HttpResponse(template.render(context, request))


# TO GET INPUTS FROM FORM FOR UPDATE BY ALLOTED MEMBERS BY CLICKING ON UNIQUE ID
@login_required
def updaterecord1(request, id):

  
  
  call_Type = request.POST['call_Type']
  account_Status=request.POST['account_Status']
  disposition = request.POST['disposition']
  denial_Code = request.POST['denial_Code']
  step = request.POST['step']
  next_Best_Action_For_Success_Rate = request.POST['next_Best_Action_For_Success_Rate']
  follow_up_Date = request.POST['follow_up_Date']
  follow_up_Comments = request.POST['follow_up_Comments']
  final_Status = request.POST['final_Status']
  status_history = request.POST['status_history']

  member = Detail.objects.get(id=id)

  previous_dis=member.disposition  # storing previously saved comment and disposition in a varuiables to use it in condition to check either one of then is now changed or not
  previous_comm=member.follow_up_Comments


  member.call_Type = call_Type
  member.account_Status = account_Status
  member.disposition = disposition
  member.denial_Code = denial_Code
  member.step = step
  member.next_Best_Action_For_Success_Rate = next_Best_Action_For_Success_Rate
  member.follow_up_Date = follow_up_Date
  member.follow_up_Comments = follow_up_Comments
  member.final_Status = final_Status

  if (previous_dis!=disposition) or (previous_comm!=follow_up_Comments):
    member.status_history =status_history+request.user.username
  member.save()
  return HttpResponseRedirect(reverse('taskalloted'))

# TO GET FORM (LEADUPDATE.HTML) TEMPLATE AFTER CLICKING UNIQUE ID 
def updatebyleader(request, id):
  mymember = Detail.objects.get(id=id)

  template = loader.get_template('leadupdate.html')
  context = {
    'mymember': mymember,
  }
  return HttpResponse(template.render(context, request))

# TO GET INPUTS FROM FORM FOR UPDATE BY LEADER BY CLICKING ON UNIQUE ID
def updaterecord(request, id):
  received = request.POST['received']
  account_or_Claim_Number = request.POST['account_or_Claim_Number']
  service_Date = request.POST['service_Date']
  payer = request.POST['payer']
  claim_Status = request.POST['claim_Status']
  charge_USD = request.POST['charge_USD']
  balance_USD = request.POST['balance_USD']
  client = request.POST['client']
  assigned_To = request.POST['assigned_To']
  upload_Date = request.POST['upload_Date']
  allocation_Date = request.POST['allocation_Date']
  member = Detail.objects.get(id=id)
  member.received = received
  member.account_or_Claim_Number = account_or_Claim_Number
  member.service_Date = service_Date
  member.payer = payer
  member.claim_Status = claim_Status
  member.charge_USD = charge_USD
  member.balance_USD = balance_USD
  member.client = client
  member.assigned_To = assigned_To
  member.upload_Date = upload_Date
  member.allocation_Date = allocation_Date
  member.save()
  return HttpResponseRedirect(reverse('topuser'))






# def customer_list(request):
#     current_user = request.user
#     userP = UserProfile.objects.get_or_create(username=current_user)
#     customer_list = Customer.objects.filter(company=userP[0].company.comp_name)
#     myFilter = TableFilter(request.GET, queryset=customer_list.all())


#     context = {
#         'customer_list': customer_list,
#         'myFilter': myFilter

#     }
#     return render(request, 'customer_list.html', context)



#to get current logged-in user
@login_required
def sample_view(request):
    current_user = request.user
    return HttpResponse(current_user)


# @login_required
# def prem(request):
    








@login_required
def testing(request):

    current_user = request.user
    mydata = Detail.objects.filter(assigned_To=current_user).order_by('final_Status').values() #to show user their alloted tasks
    
    template = loader.get_template('template1.html')
    context = {
    'mymembers': mydata,
        }
    return HttpResponse(template.render(context, request))
    #mydata = Detail.objects.all()
    #mydata = Member.objects.values_list('firstname') #To get a particular column
    #mydata = Member.objects.filter(final_Status='Completed',assigned_To='Imtiaz').values() #To filter rows based on value
    #mydata = Member.objects.filter(Q(firstname='Emil') | Q(firstname='Tobias')).values()  #for either condition
    #.filter(firstname__startswith='L');  #for value starts with
  



    




# uploading file
@login_required        #Restrict access to logged in users in Function based views
def my_view(request):
    form = FileUploadForm(request.POST, request.FILES)
    if form.is_valid():
        form.save()
    else:
        print("error")
        # display errors
    return render(request, 'index.html', {"form":form}) 



#home page with navigation bars

class HomePageView(TemplateView):
    
    template_name='nav.html'




# update fields form using modelform method
@login_required
def UpdateFields(request):
    form= UpdateDetailsForm(request.POST or None)
    if form.is_valid():
        form.save()
  
    context= {'form': form }
        
    return render(request, 'form.html', context)


# resource assign form using modelform method
@login_required
def showform(request):
    form= AssignForm(request.POST or None)
    if form.is_valid():
        form.save()
  
    context= {'form': form }
        
    return render(request, 'form.html', context)

# for viewing Resource assignment model as a table
@login_required
def record1(request):
    tasks = ResourceAssignment.objects.all()
    return render(request, 'list.html', locals())






@login_required
def Resource(request):


     
    assign = ResourceAssignmentForm()  
    return render(request,"form.html",{'form':assign})  

    #context={}

    # Res_form=ResourceAssignmentForm(request.POST or None,request.FILES or None)

    # if Res_form.is_valid():

    #     Res_form.save()

    # context['Res_form']=Res_form
    # return render(request,"vrform.html",context)


@login_required
def importer(request):
    df=pd.read_excel(r"C:\Users\Lenovo\Desktop\Django exercises\companyproject\website\ardailytracker\static\upload\ARDAILYTRACKER.xlsx")
    df.columns=["Received", "Claim", "Service Date", "Payer","Status","Charge $","Balance $","Comments","Calling \ Non Calling","Disposition","Denial Code","Step","Worked by","Worked on","Follow up Date"]
    #df.set_index("id", inplace=True)
    print(df.dtypes)
    user = settings.DATABASES['default']['USER']
    password = settings.DATABASES['default']['PASSWORD']
    database_name = settings.DATABASES['default']['NAME']

    database_url = 'postgresql://{user}:{password}@localhost:3306/{database_name}'.format(
        user=root,
        password=sqlite3,
        database_name=db.sqlite3,
    )
    engine = create_engine(database_url, echo=False)
    df.to_sql(Detail, con=engine,  if_exists='append')


# def index(request):  
#     if request.method == 'POST':  
#         task = Det_Form(request.POST, request.FILES)  
#         if task.is_valid():  
#             handle_uploaded_file(request.FILES['file'])  
#             return HttpResponse("File uploaded successfully")  
#     else:  
#         task = Det_Form()  
#         return render(request,"index.html",{'form':task})

@login_required
def index(request):
    if "GET" == request.method:
        return render(request, 'index.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["HNS"]
        print(worksheet)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)

        return render(request, 'index.html', {"excel_data":excel_data})  


# TO UPLOAD EXCEL FILE IN DATABASE
@login_required
def Import_Excel_pandas(request):
    
    if request.method == 'POST' and request.FILES['myfile']:      
        myfile = request.FILES['myfile']
        fs = FileSystemStorage()
        filename = fs.save(myfile.name, myfile)
        uploaded_file_url = fs.url(filename)              
        empexceldata = pd.read_excel(filename)        
        dbframe = empexceldata
        #dbframe["Service Date"]=pd.to_datetime(dbframe["Service Date"],format='%m-%d-%Y')
        dbframe["Service Date"]=dbframe["Service Date"].astype('datetime64[ns]').dt.date
        dbframe.rename(columns={'Follow up Date':'FollowUpDate','Worked on':'WorkedOn','Worked by':'WorkedBy','Charge $':'charge','Balance $':'balance','Service Date':'Service_Date','Calling \ Non Calling':'call_type','Denial Code':'DenialCode'},inplace=True)
        for dbframe in dbframe.itertuples():
            obj = Detail.objects.create(received=dbframe.Received,account_or_Claim_Number=dbframe.Claim, 
                                        payer=dbframe.Payer, claim_Status=dbframe.Status, service_Date=dbframe.Service_Date,
                                        charge_USD=dbframe.charge, balance_USD=dbframe.balance)           
              
            #comment_1=dbframe.Comments,call_Type=dbframe.call_type,disposition=dbframe.Disposition,denial_Code=dbframe.DenialCode,step=dbframe.Step,follow_up_Date=dbframe.FollowUpDate, worked_By=dbframe.WorkedBy, worked_On=dbframe.WorkedOn
            
            
            obj.save()
        return render(request, 'Import_excel_db.html', {
            'uploaded_file_url': uploaded_file_url
        })   
    return render(request, 'Import_excel_db.html',{})




#Importing data from excel into database using  import-export library


from tablib import Dataset
from .resources import DetailResource

@login_required
def Import_excel(request):
    if request.method == 'POST' :
        Detail =DetailResource()
        dataset = Dataset()
        new_detail = request.FILES['myfile']
        data_import = dataset.load(new_detail.read())
        result = DetailResource.import_data(dataset,dry_run=True)
        if not result.has_errors():
            DetailResource.import_data(dataset,dry_run=False)        
    return render(request, 'Import_excel_db.html',{})







@login_required
def add_details(request):
    context={}


    form=DetailForm(request.POST or None, request.FILES or None)


    if form.is_valid():
        form.save()

    context['form']=form

    return render(request,"create.html",context)






















# , UploadFileForm
# from django.http import HttpResponseRedirect
# # Imaginary function to handle an uploaded file.
# from somewhere import handle_uploaded_file

# def upload_file(request):
#     if request.method == 'POST':
#         form = UploadFileForm(request.POST, request.FILES)
#         if form.is_valid():
#             handle_uploaded_file(request.FILES['file'])
#             return HttpResponseRedirect('/success/url/')
#     else:
#         form = UploadFileForm()
#     return render(request, 'upload.html', {'form': form})













# def add_details(request):

#     context
#         if request.method == 'POST': #to make sure that the user has clicked the 'Post' button on the template file
#             if request.POST.get('received') and request.POST.get('claim') and request.POST.get('service_date') and request.POST.get('payers') and request.POST.get('status') and request.POST.get('comments') and request.POST.get('call_type') and request.POST.get('disposition') and request.POST.get('step') and request.POST.get('worked_by') and request.POST.get('worked_on'): #to check theres no blank value in these fields
#                 detail=Detail() #sets the variable equal to the Detail database
#                 detail.received= request.POST.get('received')
#                 detail.claim= request.POST.get('claim')
#                 detail.service_date= request.POST.get('service_date')
#                 detail.payers= request.POST.get('payers')
#                 detail.status= request.POST.get('status')
#                 detail.comments= request.POST.get('comments')
#                 detail.call_type= request.POST.get('call_type')
#                 detail.disposition= request.POST.get('disposition')
#                 detail.denial_code= request.POST.get('denial_code')
#                 detail.step= request.POST.get('step')
#                 detail.worked_by= request.POST.get('worked_by')
#                 detail.worked_on= request.POST.get('worked_on')
#                 detail.follow_up_date= request.POST.get('follow_up_date')
#                 detail.save()
                
#                 return render(request, 'posts/create.html')  

#         else:
#                 return render(request,'posts/create.html')